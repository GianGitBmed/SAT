# ============================================================================
# ciometro.py
# Frontend modulo CIometro
# ============================================================================

import flet as ft
import tkinter as tk
from tkinter import filedialog

import utils
import sat_service

import threading
import os
import openpyxl
import config

# ============================================================================
# SEZIONE 2 - STILE GLOBALE
# ============================================================================
INPUT_STYLE = {
    "width": 600,
    "dense": True,
    "filled": True,
    "border": ft.InputBorder.OUTLINE,
    "border_radius": 8,
    "border_color": ft.Colors.GREY_700,
    "focused_border_color": ft.Colors.BLUE,
    "bgcolor": ft.Colors.GREY_900,
}

# ============================================================================
# SEZIONE 3 - WIDGET SELECT SINGOLA
# ============================================================================
class SingleSelectWidget(ft.Column):
    def __init__(self, options, label, on_dirty=None):
        super().__init__(spacing=0)
        self.options = options
        self.selected = ""
        self.on_dirty = on_dirty

        self.search_field = ft.TextField(
            label=label,
            on_change=self.filter_options,
            on_submit=self.on_submit,
            **INPUT_STYLE
        )

        self.suggestions_list = ft.ListView(spacing=0, height=150)

        self.suggestions_card = ft.Card(
            content=self.suggestions_list,
            visible=False,
            elevation=10,
            bgcolor=ft.Colors.GREY_800,
            margin=ft.Margin.only(top=5)
        )

        self.controls = [self.search_field, self.suggestions_card]

    def filter_options(self, e):
        q = (self.search_field.value or "").lower().strip()
        self.suggestions_list.controls.clear()

        if self.on_dirty:
            self.on_dirty()

        if q and q != self.selected.lower().strip():
            matches = [o for o in self.options if q in o.lower()][:20]
            for m in matches:
                self.suggestions_list.controls.append(
                    ft.ListTile(
                        title=ft.Text(m, size=14, color=ft.Colors.WHITE),
                        on_click=lambda e, val=m: self.select_item(val),
                        hover_color=ft.Colors.GREY_700
                    )
                )
            self.suggestions_card.visible = len(self.suggestions_list.controls) > 0
        else:
            self.suggestions_card.visible = False

        self.update()

    def select_item(self, val):
        self.selected = val
        self.search_field.value = val
        self.suggestions_card.visible = False
        self.update()

        if self.on_dirty:
            self.on_dirty()

    def on_submit(self, e):
        if self.suggestions_card.visible and self.suggestions_list.controls:
            self.select_item(self.suggestions_list.controls[0].title.value)
        else:
            if self.on_dirty:
                self.on_dirty()

    def get_value(self):
        return self.selected

    def clear(self):
        self.selected = ""
        self.search_field.value = ""
        self.suggestions_card.visible = False
        self.update()

    def is_valid(self):
        return bool(self.selected)

# ============================================================================
# SEZIONE 4 - WIDGET SELECT MULTIPLA
# ============================================================================
class MultiSelectWidget(ft.Column):
    def __init__(self, options, label, on_dirty=None):
        super().__init__(spacing=5)
        self.options = options
        self.selected = []
        self.on_dirty = on_dirty

        self.search_field = ft.TextField(
            label=label,
            on_change=self.filter_options,
            on_submit=self.on_submit,
            **INPUT_STYLE
        )

        self.suggestions_list = ft.ListView(spacing=0, height=150)

        self.suggestions_card = ft.Card(
            content=self.suggestions_list,
            visible=False,
            elevation=10,
            bgcolor=ft.Colors.GREY_800,
            margin=ft.Margin.only(top=5)
        )

        self.chips_row = ft.Row(wrap=True, width=600)

        self.controls = [self.search_field, self.suggestions_card, self.chips_row]

    def filter_options(self, e):
        q = (self.search_field.value or "").lower().strip()
        self.suggestions_list.controls.clear()

        if self.on_dirty:
            self.on_dirty()

        if q:
            matches = [o for o in self.options if q in o.lower() and o not in self.selected][:20]
            for m in matches:
                self.suggestions_list.controls.append(
                    ft.ListTile(
                        title=ft.Text(m, size=14, color=ft.Colors.WHITE),
                        on_click=lambda e, val=m: self.add_item(val),
                        hover_color=ft.Colors.GREY_700
                    )
                )
            self.suggestions_card.visible = len(self.suggestions_list.controls) > 0
        else:
            self.suggestions_card.visible = False

        self.update()

    def add_item(self, val):
        if val not in self.selected:
            self.selected.append(val)

        self.search_field.value = ""
        self.suggestions_card.visible = False
        self.render_chips()

        if self.on_dirty:
            self.on_dirty()

    def on_submit(self, e):
        if self.suggestions_card.visible and self.suggestions_list.controls:
            self.add_item(self.suggestions_list.controls[0].title.value)
        else:
            if self.on_dirty:
                self.on_dirty()

    def render_chips(self):
        self.chips_row.controls.clear()

        for s in self.selected:
            display_text = s.split(" - ")[-1] if " - " in s else s
            self.chips_row.controls.append(
                ft.Chip(
                    label=ft.Text(
                        display_text,
                        color=ft.Colors.WHITE,
                        weight=ft.FontWeight.W_600
                    ),
                    bgcolor=ft.Colors.BLUE_700,
                    on_delete=lambda e, val=s: self.remove_item(val)
                )
            )

        self.update()

    def remove_item(self, val):
        if val in self.selected:
            self.selected.remove(val)
        self.render_chips()

        if self.on_dirty:
            self.on_dirty()

    def get_selected_values(self):
        return self.selected

    def clear(self):
        self.selected = []
        self.search_field.value = ""
        self.suggestions_card.visible = False
        self.render_chips()
        self.update()

    def is_valid(self):
        return len(self.selected) > 0

# ============================================================================
# SEZIONE 5 - VIEW PRINCIPALE
# ============================================================================
class CiometroView(ft.Container):
    def __init__(self, app_context):
        super().__init__(expand=True, padding=30)  # <--- Risolve il problema
        self.app = app_context
        self.form_validated = False

        self.db_sd = utils.get_db_list(self.app.local_db, "solution_designs")
        self.db_domains = utils.get_db_list(self.app.local_db, "domains")
        self.db_teams = utils.get_db_list(self.app.local_db, "teams")
        self.db_offices = utils.get_db_list(self.app.local_db, "offices")
        self.db_bb = utils.get_db_list(self.app.local_db, "bb_instances")
        self.db_app_modules = utils.get_db_list(self.app.local_db, "app_modules")
        self.db_tech = utils.get_db_list(self.app.local_db, "technologies")

        self.build_ui()
        self.refresh_buttons_state()

    # =========================================================================
    # SEZIONE 6 - COSTRUZIONE UI (TABS & LAYOUT)
    # =========================================================================
    def build_ui(self):
        self.entry_name = ft.TextField(label="Nome CI", on_change=self.mark_form_dirty, **INPUT_STYLE)
        self.entry_desc = ft.TextField(label="Descrizione", on_change=self.mark_form_dirty, **INPUT_STYLE)

        self.widget_sd = SingleSelectWidget(self.db_sd, "Solution Design", on_dirty=self.mark_form_dirty)
        self.widget_domains = MultiSelectWidget(self.db_domains, "Domini", on_dirty=self.mark_form_dirty)
        self.widget_maint_team = SingleSelectWidget(self.db_teams, "Maintenance Dev Team", on_dirty=self.mark_form_dirty)
        self.widget_change_teams = MultiSelectWidget(self.db_teams, "Change Dev Teams", on_dirty=self.mark_form_dirty)
        self.widget_maint_office = SingleSelectWidget(self.db_offices, "Maintenance ICT Office", on_dirty=self.mark_form_dirty)
        self.widget_change_offices = MultiSelectWidget(self.db_offices, "Change ICT Offices", on_dirty=self.mark_form_dirty)
        self.widget_bb = SingleSelectWidget(self.db_bb, "Building Block", on_dirty=self.mark_form_dirty)
        self.widget_app_modules = MultiSelectWidget(self.db_app_modules, "Application Modules", on_dirty=self.mark_form_dirty)
        self.widget_tech = SingleSelectWidget(self.db_tech, "Tecnologia", on_dirty=self.mark_form_dirty)

        self.txt_status = ft.Text("Compila tutti i campi e premi 'Valida inserimento'.", size=13, color=ft.Colors.GREY_400)

        # Pulsanti
        self.btn_validate = ft.ElevatedButton(
            "Valida inserimento",
            icon=ft.Icons.FACT_CHECK,
            on_click=self.validate_form,
            style=ft.ButtonStyle(bgcolor=ft.Colors.AMBER_700, color=ft.Colors.WHITE, padding=20, shape=ft.RoundedRectangleBorder(radius=8))
        )

        self.btn_submit = ft.ElevatedButton(
            "Inserisci CI",
            icon=ft.Icons.SAVE,
            disabled=True,
            on_click=self.execute_single_creation,
            style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_700, color=ft.Colors.WHITE, padding=20, shape=ft.RoundedRectangleBorder(radius=8))
        )

        self.btn_template = ft.OutlinedButton(
            "Scarica Template",
            icon=ft.Icons.FILE_DOWNLOAD,
            on_click=self.download_template,
            style=ft.ButtonStyle(color=ft.Colors.BLUE_300, padding=20, shape=ft.RoundedRectangleBorder(radius=8))
        )

        self.btn_clear = ft.OutlinedButton(
            "Pulisci",
            icon=ft.Icons.CLEANING_SERVICES,
            on_click=self.clear_form,
            style=ft.ButtonStyle(color=ft.Colors.RED_300, padding=20, shape=ft.RoundedRectangleBorder(radius=8))
        )

        # --- LAYOUT A DUE COLONNE (Il tuo originale intatto) ---
        colonna_sinistra = ft.Column(
            expand=True, spacing=15,
            controls=[self.entry_name, self.entry_desc, self.widget_sd, self.widget_maint_team, self.widget_change_teams, self.widget_bb]
        )

        colonna_destra = ft.Column(
            expand=True, spacing=15,
            controls=[self.widget_domains, self.widget_maint_office, self.widget_change_offices, self.widget_app_modules, self.widget_tech]
        )

        form_layout = ft.Row(
            controls=[colonna_sinistra, colonna_destra],
            alignment=ft.MainAxisAlignment.START,
            vertical_alignment=ft.CrossAxisAlignment.START,
            spacing=40,
        )

        # =========================================================
        # WRAPPER FINALE IN SCHEDE (CUSTOM TABS BULLETPROOF)
        # =========================================================
        
        # Riga bottoni per il Tab 1
        btn_row = ft.Row(
            controls=[self.btn_validate, self.btn_submit, self.btn_template, self.btn_clear],
            alignment=ft.MainAxisAlignment.END
        )

        # TAB 1: Inserimento Singolo (Salvato come attributo di classe per poterlo nascondere)
        self.form_singolo_content = ft.Column(
            expand=True,
            scroll=ft.ScrollMode.AUTO,
            visible=True, # Inizialmente visibile
            controls=[
                btn_row, 
                ft.Divider(height=30, color=ft.Colors.GREY_800),
                form_layout,  # Il layout a due colonne
                ft.Container(height=40),
                self.txt_status
            ]
        )

        # TAB 2: Inserimento Massivo ()
        # =========================================================
        # INSERIMENTO MASSIVO
        # =========================================================
        # --- INIT COMPONENTI MASSIVO ---
        self.massivo_selected_filepath = None

        self.txt_massivo_file = ft.Text("Nessun file selezionato", color=ft.Colors.GREY_500, italic=True)
        
        self.btn_massivo_pick = ft.ElevatedButton(
            "Seleziona Excel Template",
            icon=ft.Icons.FOLDER_OPEN,
            on_click=self.pick_file_native,
            style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_GREY_700, color=ft.Colors.WHITE)
        )
        
        self.btn_massivo_start = ft.ElevatedButton(
            "Avvia Censimento Massivo",
            icon=ft.Icons.PLAY_ARROW,
            on_click=self.start_massive_import,
            disabled=True,
            style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE)
        )

        # NUOVO BOTTONE: Anteprima
        self.btn_massivo_preview = ft.ElevatedButton(
            "Anteprima Dati",
            icon=ft.Icons.TABLE_VIEW,
            on_click=self.show_preview_dialog,
            disabled=True,
            style=ft.ButtonStyle(bgcolor=ft.Colors.INDIGO_700, color=ft.Colors.WHITE)
        )
        
        self.massivo_progress = ft.ProgressBar(width=600, value=0, visible=False, color=ft.Colors.GREEN_400)
        self.txt_massivo_status = ft.Text("Pronto.", color=ft.Colors.AMBER_300, size=14)
        
        # ELIMINATO self.massivo_log_list e relativo Container

        # TAB 2: Inserimento Massivo
        self.form_massivo_content = ft.Column(
            expand=True,
            alignment=ft.MainAxisAlignment.START,
            visible=False,
            controls=[
                ft.Text("Censimento Massivo CI", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.PURPLE_300),
                ft.Text("Importa il template Excel per creare multipli Configuration Item in un singolo job.", color=ft.Colors.GREY_400),
                ft.Divider(height=20, color=ft.Colors.GREY_800),
                
                ft.Row([self.btn_massivo_pick, self.txt_massivo_file], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Container(height=10),
                # Aggiunto il bottone di preview accanto allo start
                ft.Row([self.btn_massivo_start, self.btn_massivo_preview, self.massivo_progress], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                self.txt_massivo_status
                # ELIMINATO il Container del log visivo
            ]
        )
        # --- LOGICA CUSTOM TABS ---
        def switch_tab(e, index):
            # Cambia i colori dei tab
            tab_singolo.bgcolor = ft.Colors.BLUE_800 if index == 0 else ft.Colors.TRANSPARENT
            tab_massivo.bgcolor = ft.Colors.BLUE_800 if index == 1 else ft.Colors.TRANSPARENT
            
            # Mostra/Nascondi i contenuti
            self.form_singolo_content.visible = (index == 0)
            self.form_massivo_content.visible = (index == 1)
            
            self.update()

# "Linguette" dei Tab costruite con Container
        tab_singolo = ft.Container(
            content=ft.Row([ft.Icon(ft.Icons.PERSON, size=20), ft.Text("Censimento Singolo", weight=ft.FontWeight.BOLD)]),
            padding=ft.Padding.symmetric(horizontal=20, vertical=10),
            bgcolor=ft.Colors.BLUE_800, # Tab Attivo
            border_radius=ft.BorderRadius.all(8),
            on_click=lambda e: switch_tab(e, 0),
            ink=True  # Rimosso il cursor
        )

        tab_massivo = ft.Container(
            content=ft.Row([ft.Icon(ft.Icons.GROUPS, size=20), ft.Text("Censimento Massivo", weight=ft.FontWeight.BOLD)]),
            padding=ft.Padding.symmetric(horizontal=20, vertical=10),
            bgcolor=ft.Colors.TRANSPARENT, # Tab Inattivo
            border_radius=ft.BorderRadius.all(8),
            on_click=lambda e: switch_tab(e, 1),
            ink=True  # Rimosso il cursor
        )

        custom_tabs_header = ft.Row([tab_singolo, tab_massivo], spacing=10)

        # 4. Assegnazione globale alla vista
        self.content = ft.Column(
            expand=True,
            controls=[
                ft.Text("Nuovo Configuration Item", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_200),
                custom_tabs_header,
                ft.Divider(height=2, color=ft.Colors.GREY_800), # Sottile linea di separazione sotto i tab
                self.form_singolo_content,
                self.form_massivo_content
            ]
        )
# =========================================================================
    # METODI TAB: CENSIMENTO MASSIVO
    # =========================================================================
   # =========================================================================
    # METODI TAB: CENSIMENTO MASSIVO
    # =========================================================================
    def pick_file_native(self, e):
        """Usa Tkinter per aprire un file dialog nativo OS per l'import"""
        root = tk.Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        
        file_path = filedialog.askopenfilename(
            title="Seleziona il Template Excel",
            filetypes=[("Excel files", "*.xlsx")]
        )
        
        root.destroy()

        if file_path:
            self.massivo_selected_filepath = file_path
            self.txt_massivo_file.value = os.path.basename(file_path)
            self.txt_massivo_file.color = ft.Colors.GREEN_300
            self.btn_massivo_start.disabled = False
            self.btn_massivo_preview.disabled = False # Abilita l'anteprima
            self.log_massivo(f"File caricato in memoria: {file_path}")
        else:
            self.log_massivo("Selezione file annullata dall'utente.", is_error=True)
            
        self.app.page.update()

    def log_massivo(self, message: str, is_error: bool = False):
        # Rimossa l'aggiunta al ListView locale, logga solo sul terminale globale in app.py
        self.app.log(f"[MASSIVO] {message}", level="ERROR" if is_error else "INFO")
        self.app.page.update()

    def _map_id_to_logical(self, id_str: str, db_key: str) -> str:
        """Converte una stringa di ID (es '123|456') nei nomi logici presi dal DB in RAM"""
        if not id_str:
            return ""
        db = self.app.local_db.get(db_key, {})
        # Splittiamo per la pipe in caso di array massivi
        ids = [i.strip() for i in str(id_str).split('|') if i.strip()]
        names = [db.get(i, f"ID:{i}") for i in ids]
        return " | ".join(names)

    def show_preview_dialog(self, e):
        """Legge l'Excel e crea un data table dinamico mappato coi nomi logici"""
        if not self.massivo_selected_filepath:
            return

        try:
            wb = openpyxl.load_workbook(self.massivo_selected_filepath, data_only=True)
            sheet = wb.active

            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            if "name" not in headers:
                self.show_dialog("Errore Anteprima", "Il file Excel non contiene la colonna 'name'.", is_error=True)
                return

            # Costruiamo le colonne della tabella Flet
            columns = [ft.DataColumn(ft.Text(h, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_200)) for h in headers]
            rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if not row_dict.get("name"):
                    continue
                
                cells = []
                for h in headers:
                    raw_val = str(row_dict.get(h) or "").strip()
                    
                    # Routing del mapping in base al nome colonna di Distinta
                    if h == "solutionDesignId": logical_val = self._map_id_to_logical(raw_val, "solution_designs")
                    elif h == "domainIds": logical_val = self._map_id_to_logical(raw_val, "domains")
                    elif h in ["maintenanceDevelopmentTeamId", "changeDevelopmentTeamIds"]: logical_val = self._map_id_to_logical(raw_val, "teams")
                    elif h in ["maintenanceIctOfficeId", "changeIctOfficeIds"]: logical_val = self._map_id_to_logical(raw_val, "offices")
                    elif h == "buildingBlockInstanceId": logical_val = self._map_id_to_logical(raw_val, "bb_instances")
                    elif h == "applicationModuleIds": logical_val = self._map_id_to_logical(raw_val, "app_modules")
                    elif h == "technologyId": logical_val = self._map_id_to_logical(raw_val, "technologies")
                    else: logical_val = raw_val # Name, Description, ecc...
                    
                    cells.append(ft.DataCell(ft.Text(logical_val, size=12)))
                
                rows.append(ft.DataRow(cells=cells))

            # Renderizziamo la tabella dentro un dialog largo
            table = ft.DataTable(
                columns=columns, 
                rows=rows, 
                heading_row_color=ft.Colors.GREY_900,
                border=ft.border.all(1, ft.Colors.GREY_800),
                vertical_lines=ft.border.BorderSide(1, ft.Colors.GREY_800)
            )

            dlg = ft.AlertDialog(
                title=ft.Row([ft.Icon(ft.Icons.TABLE_CHART, color=ft.Colors.BLUE), ft.Text("Anteprima Dati (Risoluzione Logica)")]),
                # Doppio scroll per supportare tabelle con molte righe/colonne
                content=ft.Container(
                    content=ft.Column([ft.Row([table], scroll=ft.ScrollMode.ALWAYS)], scroll=ft.ScrollMode.ALWAYS),
                    width=1200, 
                    height=600,
                    border_radius=10
                ),
                actions=[ft.TextButton("Chiudi", on_click=lambda e: self.close_dialog(dlg))]
            )
            self.app.page.show_dialog(dlg)

        except Exception as ex:
            self.app.log(f"Errore generazione anteprima: {str(ex)}", level="ERROR")
            self.show_dialog("Errore Anteprima", f"Impossibile leggere il file:\n{str(ex)}", is_error=True)
            
    def start_massive_import(self, e):
        if not self.massivo_selected_filepath:
            return
            
        api_key = self.app.config.get("api_key", "")
        if not api_key:
            self.log_massivo("ERRORE CRITICO: API Key mancante. Vai in Impostazioni per configurarla.", is_error=True)
            return

        # Lock UI controls
        self.btn_massivo_start.disabled = True
        self.btn_massivo_pick.disabled = True
        self.massivo_progress.value = 0
        self.massivo_progress.visible = True
        self.txt_massivo_status.value = "Lettura del file Excel in corso..."
        self.app.page.update()

        # Fire and forget nel thread
        threading.Thread(target=self._process_excel_thread, args=(api_key,), daemon=True).start()

    def _process_excel_thread(self, api_key: str):
        try:
            wb = openpyxl.load_workbook(self.massivo_selected_filepath, data_only=True)
            sheet = wb.active

            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            if "name" not in headers:
                self.log_massivo("ERRORE: Il file Excel non contiene la colonna 'name'. Template non valido.", is_error=True)
                self._unlock_massivo_ui("Operazione annullata. Template errato.")
                return

            valid_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if not row_dict.get("name"):
                    continue
                    
                # Safe stringify mapping coerente col sat_service
                stringified_dict = {k: str(v).strip() if v is not None else "" for k, v in row_dict.items() if k}
                valid_rows.append(stringified_dict)

            total_cis = len(valid_rows)
            if total_cis == 0:
                self.log_massivo("Nessun dato valido trovato nel file Excel.", is_error=True)
                self._unlock_massivo_ui("File vuoto.")
                return

            self.log_massivo(f"Inizio elaborazione: {total_cis} CI rilevati.")
            self.txt_massivo_status.value = "Censimento API in corso..."
            
            success_count = 0
            error_count = 0

            for i, snapshot in enumerate(valid_rows):
                ci_name = snapshot.get("name", "Sconosciuto")
                self.log_massivo(f"[{i+1}/{total_cis}] Invio: {ci_name} ...")
                
                res = sat_service.create_ci(api_key, snapshot)
                
                if res.get("success"):
                    self.log_massivo(f" ✅ OK: {ci_name} creato correttamente.")
                    success_count += 1
                else:
                    err_msg = res.get("message", "Errore API")
                    details = " | ".join(res.get("errors", []))
                    self.log_massivo(f" ❌ KO: {ci_name} - {err_msg} -> {details}", is_error=True)
                    error_count += 1

                # Progress Bar update
                self.massivo_progress.value = (i + 1) / total_cis
                self.app.page.update()

            # Aggiornamento stats globali per la Panoramica
            current_stats = self.app.config.get("stats", {})
            current_stats["ci_massivi_ok"] = current_stats.get("ci_massivi_ok", 0) + success_count
            current_stats["ci_ko"] = current_stats.get("ci_ko", 0) + error_count
            self.app.config["stats"] = current_stats
            config.save_app_config(self.app.config)

            final_msg = f"Completato. Successi: {success_count}, Errori: {error_count}"
            self.log_massivo(final_msg)
            self._unlock_massivo_ui(final_msg)

        except Exception as e:
            self.log_massivo(f"ERRORE CRITICO THREAD: {str(e)}", is_error=True)
            self._unlock_massivo_ui("Interrotto da errore di sistema.")

    def _unlock_massivo_ui(self, status_msg: str):
        self.btn_massivo_start.disabled = False
        self.btn_massivo_pick.disabled = False
        self.btn_massivo_preview.disabled = False
        self.massivo_progress.visible = False
        self.txt_massivo_status.value = status_msg
        self.app.page.update()
    # =========================================================================
    # SEZIONE 7 - LETTURA FORM
    # =========================================================================
    def collect_form_data(self) -> dict:
        return {
            "name": self.entry_name.value,
            "description": self.entry_desc.value,
            "solution_design": self.widget_sd.get_value(),
            "domains": self.widget_domains.get_selected_values(),
            "maintenance_team": self.widget_maint_team.get_value(),
            "change_teams": self.widget_change_teams.get_selected_values(),
            "maintenance_office": self.widget_maint_office.get_value(),
            "change_offices": self.widget_change_offices.get_selected_values(),
            "building_block": self.widget_bb.get_value(),
            "application_modules": self.widget_app_modules.get_selected_values(),
            "technology": self.widget_tech.get_value(),
        }

    # =========================================================================
    # SEZIONE 8 - STATO / VALIDAZIONE
    # =========================================================================
    def mark_form_dirty(self, e=None):
        self.form_validated = False
        self.txt_status.value = "Modifiche rilevate. Premi 'Valida inserimento' per riabilitare l'inserimento."
        self.txt_status.color = ft.Colors.AMBER_300
        self.refresh_buttons_state()
        self.update()

    def refresh_buttons_state(self):
        snapshot = sat_service.build_snapshot_from_form(self.collect_form_data())
        missing = sat_service.validate_snapshot(snapshot)
        enabled = (len(missing) == 0 and self.form_validated)

        self.btn_submit.disabled = not enabled
        self.btn_submit.style = ft.ButtonStyle(
            bgcolor=ft.Colors.GREEN_700 if enabled else ft.Colors.GREY_700,
            color=ft.Colors.WHITE,
            padding=20,
            shape=ft.RoundedRectangleBorder(radius=8)
        )

    def validate_form(self, e):
        self.app.log("Avvio validazione form (Check locale)...")
        snapshot = sat_service.build_snapshot_from_form(self.collect_form_data())

        # 1. CONTROLLO CAMPI MANCANTI
        missing = sat_service.validate_snapshot(snapshot)
        if missing:
            self.form_validated = False
            self.refresh_buttons_state()
            self.txt_status.value = "Validazione fallita."
            self.txt_status.color = ft.Colors.RED_300
            self.update()
            self.app.log(f"Validazione bloccata: mancano {len(missing)} campi obbligatori.", level="WARN")
            self.show_dialog("Validazione Fallita", "Compila i seguenti campi obbligatori:\n- " + "\n- ".join(missing), is_error=True)
            return

        # 2. CONTROLLO OMONIMIA IN RAM
        ci_name = snapshot.get("name", "").strip()
        ci_name_lower = ci_name.lower()
        
        existing_cis = self.app.local_db.get("configuration_items", {})
        if any(name.strip().lower() == ci_name_lower for name in existing_cis.values()):
            self.form_validated = False
            self.refresh_buttons_state()
            self.txt_status.value = "Validazione fallita: Omonimia rilevata nel DB locale."
            self.txt_status.color = ft.Colors.RED_300
            self.update()
            self.app.log(f"Validazione bloccata (Cache locale): Il CI '{ci_name}' esiste già.", level="WARN")
            self.show_dialog(
                "CI Già Esistente", 
                f"Attenzione: un Configuration Item chiamato '{ci_name}' risulta già presente nel database locale.\nCambia nome.", 
                is_error=True
            )
            return

        # 3. VALIDAZIONE SUPERATA
        self.form_validated = True
        self.refresh_buttons_state()
        self.txt_status.value = "Validazione completata. Pronto per l'invio a Distinta."
        self.txt_status.color = ft.Colors.GREEN_300
        self.update()
        
        self.app.log("Validazione superata: form completo. In attesa di invio a Distinta.")
        self.show_dialog(
            "Validazione OK", 
            "I campi sono corretti.\nPremi 'Inserisci CI' per inviare la richiesta a Distinta."
        )

    # =========================================================================
    # SEZIONE 9 - INSERIMENTO CI (NATIVE FLET THREADING)
    # =========================================================================
    def execute_single_creation(self, e):
        self.app.log("Click su Inserisci CI ricevuto.")
        
        snapshot = sat_service.build_snapshot_from_form(self.collect_form_data())
        api_key = (self.app.config.get("api_key", "") or "").strip()

        self.btn_submit.disabled = True
        self.btn_submit.text = "Inviando..."
        self.update()

        try:
            self.app.log(f"Avvio chiamata API per CI: '{snapshot.get('name')}'")
            
            result = sat_service.create_ci(api_key, snapshot)
            
            self.btn_submit.text = "Inserisci CI"
            
            if result.get("success"):
                self.txt_status.value = "Inserimento completato."
                self.txt_status.color = ft.Colors.GREEN_300
                success_msg = result.get("message", "CI creato correttamente a sistema.")
                
                self.app.log(f"SUCCESSO: {success_msg}")
                
                import config
                self.app.config.setdefault("stats", {})["ci_singoli_ok"] = self.app.config.get("stats", {}).get("ci_singoli_ok", 0) + 1
                config.save_app_config(self.app.config)

                self.refresh_buttons_state()
                self.update()
                self.show_dialog("Inserimento Riuscito", success_msg)
                
            else:
                errors = result.get("errors", [])
                msg = "\n".join([f"- {x}" for x in errors]) if errors else "Errore API imprevisto."
                
                self.txt_status.value = "Inserimento fallito."
                self.txt_status.color = ft.Colors.RED_300
                
                self.app.log(f"ERRORE API:\n{msg}", level="ERROR")
                
                import config
                self.app.config.setdefault("stats", {})["ci_ko"] = self.app.config.get("stats", {}).get("ci_ko", 0) + 1
                config.save_app_config(self.app.config)

                self.refresh_buttons_state()
                self.update()
                self.show_dialog("Errore Inserimento", msg, is_error=True)

        except Exception as ex:
            self.btn_submit.text = "Inserisci CI"
            self.app.log(f"CRASH APPLICATIVO: {str(ex)}", level="ERROR")
            
            import config
            self.app.config.setdefault("stats", {})["ci_ko"] = self.app.config.get("stats", {}).get("ci_ko", 0) + 1
            config.save_app_config(self.app.config)
            
            self.refresh_buttons_state()
            self.update()
            self.show_dialog("Errore Grave", f"Si è verificato un problema:\n{str(ex)}", is_error=True)

    # =========================================================================
    # SEZIONE 10 - EXPORT TEMPLATE
    # =========================================================================
    def download_template(self, e):
        self.app.log("Richiesta export template massivo...")
        try:
            snapshot = sat_service.build_snapshot_from_form(self.collect_form_data())

            if sat_service.is_snapshot_empty(snapshot):
                self.app.log("Download bloccato: form completamente vuoto.", level="WARN")
                self.show_dialog("Download inibito", "Compila almeno un campo del form prima di scaricare il template.", is_error=True)
                return

            root = tk.Tk()
            root.withdraw()
            root.attributes("-topmost", True)

            path = filedialog.asksaveasfilename(
                title="Salva Template Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Template_Censimento_Compilato.xlsx"
            )

            root.destroy()

            if not path:
                self.app.log("Salvataggio Excel annullato dall'utente.")
                return

            if not path.lower().endswith(".xlsx"):
                path += ".xlsx"

            sat_service.export_snapshot_to_excel(path, snapshot, "create")
            self.app.log(f"Template Excel generato con successo in: {path}")
            self.show_dialog("Successo", f"Template Excel generato correttamente in:\n{path}")

        except Exception as ex:
            self.app.log(f"Errore Generazione Excel: {str(ex)}", level="ERROR")
            self.show_dialog("Errore Generazione Excel", str(ex), is_error=True)

    # =========================================================================
    # SEZIONE 11 - PULIZIA FORM
    # =========================================================================
    def clear_form(self, e):
        self.app.log("Svuotamento dei campi del form in corso.")
        self.entry_name.value = ""
        self.entry_desc.value = ""

        self.widget_sd.clear()
        self.widget_domains.clear()
        self.widget_maint_team.clear()
        self.widget_change_teams.clear()
        self.widget_maint_office.clear()
        self.widget_change_offices.clear()
        self.widget_bb.clear()
        self.widget_app_modules.clear()
        self.widget_tech.clear()

        self.form_validated = False
        self.txt_status.value = "Form pulito. Compila i campi e valida l'inserimento."
        self.txt_status.color = ft.Colors.GREY_400

        self.btn_submit.text = "Inserisci CI"
        self.refresh_buttons_state()
        self.update()

    # =========================================================================
    # SEZIONE 12 - DIALOG (FLET 0.84+ API)
    # =========================================================================
    def show_dialog(self, title, content, is_error=False):
        icon_color = ft.Colors.RED_ACCENT if is_error else ft.Colors.BLUE_400

        dlg = ft.AlertDialog(
            title=ft.Row(
                controls=[
                    ft.Icon(ft.Icons.INFO_OUTLINE, color=icon_color),
                    ft.Text(title, weight=ft.FontWeight.BOLD),
                ]
            ),
            content=ft.Column(
                [ft.Text(content, size=14)],
                scroll=ft.ScrollMode.AUTO,
                height=200 if len(content) > 150 else None,
                tight=True
            ),
            actions=[
                ft.TextButton(
                    "Chiudi",
                    on_click=lambda e: self.close_dialog(dlg)
                )
            ],
            shape=ft.RoundedRectangleBorder(radius=10),
            bgcolor=ft.Colors.GREY_900
        )

        self.app.page.show_dialog(dlg)

    def close_dialog(self, dlg):
        if hasattr(self.app.page, "close_dialog"):
            self.app.page.close_dialog()
        elif hasattr(self.app.page, "pop_dialog"):
            self.app.page.pop_dialog()
        else:
            dlg.open = False
            self.app.page.update()
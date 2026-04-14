# excel_handler.py
from openpyxl import Workbook, load_workbook
import api_queries

def generate_template(filepath: str, template_type: str) -> bool:
    """Genera un template Excel in base al tipo richiesto ('create' o 'assoc')."""
    try:
        wb = Workbook()
        ws = wb.active
        
        if template_type == "create":
            ws.title = "Censimento_Massivo"
            ws.append(api_queries.EXCEL_CREATE_HEADER)
            for row in api_queries.EXCEL_CREATE_EXAMPLE:
                ws.append(row)
        elif template_type == "assoc":
            ws.title = "Associazione_Massiva"
            ws.append(api_queries.EXCEL_ASSOC_HEADER)
            for row in api_queries.EXCEL_ASSOC_EXAMPLE:
                ws.append(row)
        else:
            raise ValueError("Tipo template non valido.")
            
        wb.save(filepath)
        return True
    except Exception as e:
        raise Exception(f"Errore generazione Excel: {e}")

def read_massive_file(filepath: str) -> list:
    """Legge il file Excel e mappa i valori usando l'header della prima riga."""
    try:
        wb = load_workbook(filepath, data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        data = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row): continue # Skip righe vuote
            row_dict = {headers[i]: str(value) if value is not None else "" for i, value in enumerate(row) if i < len(headers)}
            data.append(row_dict)
            
        return data
    except Exception as e:
        raise Exception(f"Errore lettura Excel: {e}")

def generate_filled_template(filepath: str, data: dict, template_type: str = "create") -> bool:
    """
    Genera un file Excel con header standard e una riga valorizzata
    con i dati snapshot del form.
    """
    try:
        wb = Workbook()
        ws = wb.active

        if template_type == "create":
            ws.title = "Censimento_Massivo"
            headers = api_queries.EXCEL_CREATE_HEADER
        elif template_type == "assoc":
            ws.title = "Associazione_Massiva"
            headers = api_queries.EXCEL_ASSOC_HEADER
        else:
            raise ValueError("Tipo template non valido.")

        ws.append(headers)
        ws.append([data.get(h, "") for h in headers])

        wb.save(filepath)
        return True

    except Exception as e:
        raise Exception(f"Errore generazione Excel valorizzato: {e}")
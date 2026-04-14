# ============================================================================
# ci_esistenti.py
# Modulo SAT per l'associazione di Configuration Item Esistenti a Solution Design
# ============================================================================

import flet as ft
import tkinter as tk
from tkinter import filedialog
import threading
import os
import openpyxl

import config
import utils
from api_client import ApiClient
import api_queries

# ============================================================================
# WIDGET LOCALI (Per incapsulamento e indipendenza da ciometro.py)
# ============================================================================
INPUT_STYLE = {
    "width": 600, "dense": True, "filled": True, "border": ft.InputBorder.OUTLINE,
    "border_radius": 8, "border_color": ft.Colors.GREY_700,
    "focused_border_color": ft.Colors.BLUE, "bgcolor": ft.Colors.GREY_900,
}

class SingleSelectWidget(ft.Column):
    def __init__(self, options, label, on_dirty=None):
        super().__init__(spacing=0)
        self.options = options
        self.selected = ""
        self.on_dirty = on_dirty
        self.search_field = ft.TextField(label=label, on_change=self.filter_options, on_submit=self.on_submit, **INPUT_STYLE)
        self.suggestions_list = ft.ListView(spacing=0, height=150)
        self.suggestions_card = ft.Card(content=self.suggestions_list, visible=False, elevation=10, bgcolor=ft.Colors.GREY_800, margin=ft.margin.only(top=5))
        self.controls = [self.search_field, self.suggestions_card]

    def filter_options(self, e):
        q = (self.search_field.value or "").lower().strip()
        self.suggestions_list.controls.clear()
        if self.on_dirty: self.on_dirty()
        if q and q != self.selected.lower().strip():
            matches = [o for o in self.options if q in o.lower()][:20]
            for m in matches:
                self.suggestions_list.controls.append(ft.ListTile(title=ft.Text(m, size=14, color=ft.Colors.WHITE), on_click=lambda e, val=m: self.select_item(val), hover_color=ft.Colors.GREY_700))
            self.suggestions_card.visible = len(self.suggestions_list.controls) > 0
        else:
            self.suggestions_card.visible = False
        self.update()

    def select_item(self, val):
        self.selected = val
        self.search_field.value = val
        self.suggestions_card.visible = False
        self.update()
        if self.on_dirty: self.on_dirty()

    def on_submit(self, e):
        if self.suggestions_card.visible and self.suggestions_list.controls:
            self.select_item(self.suggestions_list.controls[0].title.value)

    def get_value(self): return self.selected
    def clear(self):
        self.selected = ""; self.search_field.value = ""; self.suggestions_card.visible = False; self.update()

class MultiSelectWidget(ft.Column):
    def __init__(self, options, label, on_dirty=None):
        super().__init__(spacing=5)
        self.options = options
        self.selected = []
        self.on_dirty = on_dirty
        self.search_field = ft.TextField(label=label, on_change=self.filter_options, on_submit=self.on_submit, **INPUT_STYLE)
        self.suggestions_list = ft.ListView(spacing=0, height=150)
        self.suggestions_card = ft.Card(content=self.suggestions_list, visible=False, elevation=10, bgcolor=ft.Colors.GREY_800, margin=ft.margin.only(top=5))
        self.chips_row = ft.Row(wrap=True, width=600)
        self.controls = [self.search_field, self.suggestions_card, self.chips_row]

    def filter_options(self, e):
        q = (self.search_field.value or "").lower().strip()
        self.suggestions_list.controls.clear()
        if self.on_dirty: self.on_dirty()
        if q:
            matches = [o for o in self.options if q in o.lower() and o not in self.selected][:20]
            for m in matches:
                self.suggestions_list.controls.append(ft.ListTile(title=ft.Text(m, size=14, color=ft.Colors.WHITE), on_click=lambda e, val=m: self.add_item(val), hover_color=ft.Colors.GREY_700))
            self.suggestions_card.visible = len(self.suggestions_list.controls) > 0
        else:
            self.suggestions_card.visible = False
        self.update()

    def add_item(self, val):
        if val not in self.selected: self.selected.append(val)
        self.search_field.value = ""; self.suggestions_card.visible = False; self.render_chips()
        if self.on_dirty: self.on_dirty()

    def on_submit(self, e):
        if self.suggestions_card.visible and self.suggestions_list.controls:
            self.add_item(self.suggestions_list.controls[0].title.value)

    def render_chips(self):
        self.chips_row.controls.clear()
        for s in self.selected:
            display_text = s.split(" - ")[-1] if " - " in s else s
            self.chips_row.controls.append(ft.Chip(label=ft.Text(display_text, color=ft.Colors.WHITE, weight=ft.FontWeight.W_600), bgcolor=ft.Colors.BLUE_700, on_delete=lambda e, val=s: self.remove_item(val)))
        self.update()

    def remove_item(self, val):
        if val in self.selected: self.selected.remove(val)
        self.render_chips()
        if self.on_dirty: self.on_dirty()

    def get_selected_values(self): return self.selected
    def clear(self):
        self.selected = []; self.search_field.value = ""; self.suggestions_card.visible = False; self.render_chips(); self.update()

# ============================================================================
# LOGICA DI SERVIZIO API (ASSOCIAZIONE & BONIFICA)
# ============================================================================
def execute_association_api(api_key: str, snapshot: dict) -> dict:
    ci_id = snapshot.get("configurationItemId")
    if not ci_id or not str(ci_id).isdigit():
        return {"success": False, "message": "ID CI mancante o non valido", "errors": []}
        
    ci_id = int(ci_id)

    # 1. RECUPERO DATI ESISTENTI DEL CI TRAMITE REST
    res_get = ApiClient.send_rest_get(api_key, f"configuration_items/{ci_id}")
    if not res_get.get("success"):
        return {"success": False, "message": f"Impossibile recuperare i dettagli del CI {ci_id} da Distinta.", "errors": [res_get.get("error", "Errore REST")]}
        
    ci_data = res_get.get("data", {})
    if "configuration_item" in ci_data:
        ci_data = ci_data["configuration_item"]

    def _extract_single(data, keys):
        for k in keys:
            val = data.get(k)
            if val is not None:
                if isinstance(val, dict): return val.get("id")
                if isinstance(val, (int, str)) and str(val).isdigit(): return int(val)
        return None

    def _extract_list(data, keys):
        for k in keys:
            val = data.get(k)
            if isinstance(val, list) and len(val) > 0:
                return [v.get("id") if isinstance(v, dict) else int(v) for v in val if v]
        return []

    # --- ESTRAZIONE DATI ESISTENTI (Potenzialmente vuoti) ---
    raw_state = ci_data.get("state", "published")
    safe_state = "draft" if raw_state == "draft" else "published"
    
    maint_team = _extract_single(ci_data, ["maintenance_development_team", "maintenance_development_team_id", "maintenanceDevelopmentTeamId"])
    change_teams = _extract_list(ci_data, ["change_development_teams", "change_development_team_ids", "changeDevelopmentTeamIds"])
    
    maint_office = _extract_single(ci_data, ["maintenance_ict_office", "maintenance_ict_office_id", "maintenanceIctOfficeId"])
    change_offices = _extract_list(ci_data, ["change_ict_offices", "change_ict_office_ids", "changeIctOfficeIds"])

    # --- LOGICA DI BONIFICA AUTOMATICA (Auto-Derivazione incrociata) ---
    if not maint_team and change_teams: maint_team = change_teams[0]
    elif not change_teams and maint_team: change_teams = [maint_team]
    
    if not maint_office and change_offices: maint_office = change_offices[0]
    elif not change_offices and maint_office: change_offices = [maint_office]

    # --- LOGICA DI BONIFICA MANUALE (Dalla UI, se i campi erano totalmente vuoti) ---
    import utils
    ui_team = utils.extract_id(snapshot.get("bonificaTeam", ""))
    if ui_team:
        maint_team = int(ui_team)
        change_teams = [int(ui_team)]
        
    ui_office = utils.extract_id(snapshot.get("bonificaOffice", ""))
    if ui_office:
        maint_office = int(ui_office)
        change_offices = [int(ui_office)]

    # 2. COSTRUZIONE PAYLOAD IBRIDO & BONIFICATO
    payload = {
        "state": safe_state,
        "environments": ["integration", "test", "preprod", "prod"],
        "type": "used",
        "warningsAccepted": False,
        "calledConfigurationItemIds": [],
        
        "configurationItemId": ci_id,
        "solutionDesignId": int(snapshot.get("solutionDesignId")) if str(snapshot.get("solutionDesignId")).isdigit() else None,
        "applicationModuleIds": [s.strip() for s in str(snapshot.get("applicationModuleIds", "")).split("|") if s.strip()],
        "description": str(snapshot.get("description", "")).strip(),
        
        "name": ci_data.get("name", ""),
        "domainIds": _extract_list(ci_data, ["domains", "domain_ids", "domainIds"]),
        "buildingBlockInstanceId": _extract_single(ci_data, ["building_block_instance", "building_block_instance_id", "buildingBlockInstanceId"]),
        "technologyId": _extract_single(ci_data, ["technology", "technology_id", "technologyId"]),
        
        # Campi processati dal motore di bonifica
        "changeDevelopmentTeamIds": change_teams,
        "maintenanceDevelopmentTeamId": maint_team,
        "changeIctOfficeIds": change_offices,
        "maintenanceIctOfficeId": maint_office
    }

    # 3. INVIO MUTATION GRAPHQL
    res = ApiClient.send_graphql(api_key, api_queries.GRAPHQL_ASSOC_CI_QUERY, {"input": payload})
    
    if not res.get("success"):
        return {"success": False, "message": "Errore di comunicazione GraphQL", "errors": [res.get("error")]}

    data = res.get("data", {}).get("createConfigurationItemNeed", {})
    if data.get("successful"):
        return {"success": True, "message": "CI associato (e bonificato ove necessario) con successo."}
        
    api_errors = [x.get("message", "Errore") for x in data.get("errors", [])]
    return {"success": False, "message": "Validazione Distinta Fallita", "errors": api_errors}

# ============================================================================
# MAIN VIEW: CI ESISTENTI
# ============================================================================
class CiEsistentiView(ft.Container):
    def __init__(self, app_context):
        super().__init__(expand=True, padding=30)
        self.app = app_context
        self.form_validated = False

        # Liste da cache locale
        self.db_ci = utils.get_db_list(self.app.local_db, "configuration_items")
        self.db_sd = utils.get_db_list(self.app.local_db, "solution_designs")
        self.db_app_modules = utils.get_db_list(self.app.local_db, "app_modules")
        # --- AGGIUNTE PER BONIFICA ---
        self.db_teams = utils.get_db_list(self.app.local_db, "teams")
        self.db_offices = utils.get_db_list(self.app.local_db, "offices")

        self.build_ui()

    def build_ui(self):
        # --- TAB 1: ASSOCIAZIONE SINGOLA ---
        self.widget_sd = SingleSelectWidget(self.db_sd, "Solution Design di destinazione", on_dirty=self.mark_form_dirty)
        self.widget_ci = SingleSelectWidget(self.db_ci, "Configuration Item Esistente", on_dirty=self.mark_form_dirty)
        self.widget_app_modules = MultiSelectWidget(self.db_app_modules, "Application Modules (Modifica)", on_dirty=self.mark_form_dirty)
        self.entry_desc = ft.TextField(label="Descrizione (Modifica)", on_change=self.mark_form_dirty, multiline=True, min_lines=3, **INPUT_STYLE)
        # --- WIDGET BONIFICA (Inizialmente Nascosti) ---
        self.widget_bonifica_team = SingleSelectWidget(self.db_teams, "🛠️ Bonifica: Team di Sviluppo", on_dirty=self.mark_form_dirty)
        self.widget_bonifica_team.visible = False
        
        self.widget_bonifica_office = SingleSelectWidget(self.db_offices, "🛠️ Bonifica: Ufficio Referente", on_dirty=self.mark_form_dirty)
        self.widget_bonifica_office.visible = False
        self.txt_status = ft.Text("Compila i campi e premi 'Valida associazione'.", size=13, color=ft.Colors.GREY_400)

        self.btn_validate = ft.ElevatedButton("Valida associazione", icon=ft.Icons.FACT_CHECK, on_click=self.validate_single, style=ft.ButtonStyle(bgcolor=ft.Colors.AMBER_700, color=ft.Colors.WHITE, padding=20))
        self.btn_submit = ft.ElevatedButton("Associa CI", icon=ft.Icons.LINK, disabled=True, on_click=self.submit_single, style=ft.ButtonStyle(bgcolor=ft.Colors.GREY_700, color=ft.Colors.WHITE, padding=20))
        self.btn_template = ft.OutlinedButton("Scarica Template", icon=ft.Icons.FILE_DOWNLOAD, on_click=self.download_template, style=ft.ButtonStyle(color=ft.Colors.BLUE_300, padding=20))

        colonna_sinistra = ft.Column(expand=True, spacing=15, controls=[self.widget_sd, self.widget_ci, self.entry_desc])
        colonna_destra = ft.Column(expand=True, spacing=15, controls=[self.widget_app_modules, self.widget_bonifica_team, self.widget_bonifica_office])
        form_layout = ft.Row(controls=[colonna_sinistra, colonna_destra], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.START, spacing=40)

        self.tab_singolo_content = ft.Column(
            expand=True, scroll=ft.ScrollMode.AUTO, visible=True,
            controls=[
                ft.Row([self.btn_validate, self.btn_submit, self.btn_template], alignment=ft.MainAxisAlignment.END), 
                ft.Divider(height=30, color=ft.Colors.GREY_800),
                form_layout, ft.Container(height=20), self.txt_status
            ]
        )

        # --- TAB 2: ASSOCIAZIONE MASSIVA ---
        self.massivo_filepath = None
        self.massivo_failed_rows = [] # <--- Lista per conservare in RAM i record falliti

        self.txt_massivo_file = ft.Text("Nessun file selezionato", color=ft.Colors.GREY_500, italic=True)
        self.btn_massivo_pick = ft.ElevatedButton("Seleziona Excel Template", icon=ft.Icons.FOLDER_OPEN, on_click=self.pick_file_native, style=ft.ButtonStyle(bgcolor=ft.Colors.BLUE_GREY_700, color=ft.Colors.WHITE))
        self.btn_massivo_start = ft.ElevatedButton("Avvia Associazione Massiva", icon=ft.Icons.PLAY_ARROW, on_click=self.start_massive, disabled=True, style=ft.ButtonStyle(bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE))
        self.btn_massivo_preview = ft.ElevatedButton("Anteprima Dati", icon=ft.Icons.TABLE_VIEW, on_click=self.show_preview_dialog, disabled=True, style=ft.ButtonStyle(bgcolor=ft.Colors.INDIGO_700, color=ft.Colors.WHITE))
        
        # <--- NUOVO BOTTONE REPORT (Inizialmente nascosto) ---
        self.btn_massivo_report = ft.OutlinedButton("Scarica Report Errori", icon=ft.Icons.FILE_DOWNLOAD, on_click=self.download_error_report, visible=False, style=ft.ButtonStyle(color=ft.Colors.RED_300))

        self.massivo_progress = ft.ProgressBar(width=600, value=0, visible=False, color=ft.Colors.GREEN_400)
        self.txt_massivo_status = ft.Text("Pronto.", color=ft.Colors.AMBER_300, size=14)

        self.tab_massivo_content = ft.Column(
            expand=True, alignment=ft.MainAxisAlignment.START, visible=False,
            controls=[
                ft.Text("Associazione Massiva CI Esistenti", size=24, weight=ft.FontWeight.BOLD, color=ft.Colors.PURPLE_300),
                ft.Text("Associa molteplici CI ai rispettivi Solution Design tramite template Excel.", color=ft.Colors.GREY_400),
                ft.Divider(height=20, color=ft.Colors.GREY_800),
                ft.Row([self.btn_massivo_pick, self.txt_massivo_file], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                ft.Container(height=10),
                # Aggiunto il bottone report alla riga dei comandi
                ft.Row([self.btn_massivo_start, self.btn_massivo_preview, self.btn_massivo_report, self.massivo_progress], alignment=ft.MainAxisAlignment.START, vertical_alignment=ft.CrossAxisAlignment.CENTER),
                self.txt_massivo_status
            ]
        )

        # --- CUSTOM TABS LOGIC ---
        def switch_tab(e, index):
            tab_s.bgcolor = ft.Colors.BLUE_800 if index == 0 else ft.Colors.TRANSPARENT
            tab_m.bgcolor = ft.Colors.BLUE_800 if index == 1 else ft.Colors.TRANSPARENT
            self.tab_singolo_content.visible = (index == 0)
            self.tab_massivo_content.visible = (index == 1)
            self.update()

        tab_s = ft.Container(content=ft.Row([ft.Icon(ft.Icons.PERSON, size=20), ft.Text("Associazione Singola", weight=ft.FontWeight.BOLD)]), padding=ft.padding.symmetric(horizontal=20, vertical=10), bgcolor=ft.Colors.BLUE_800, border_radius=8, on_click=lambda e: switch_tab(e, 0))
        tab_m = ft.Container(content=ft.Row([ft.Icon(ft.Icons.GROUPS, size=20), ft.Text("Associazione Massiva", weight=ft.FontWeight.BOLD)]), padding=ft.padding.symmetric(horizontal=20, vertical=10), bgcolor=ft.Colors.TRANSPARENT, border_radius=8, on_click=lambda e: switch_tab(e, 1))

        self.content = ft.Column(expand=True, controls=[
            ft.Text("Aggiungi Configuration Item Esistente", size=28, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_200),
            ft.Row([tab_s, tab_m], spacing=10), ft.Divider(height=2, color=ft.Colors.GREY_800),
            self.tab_singolo_content, self.tab_massivo_content
        ])

    # =========================================================================
    # LOGICA SINGOLO
    # =========================================================================
    def get_single_snapshot(self):
        return {
            "solutionDesignId": utils.extract_id(self.widget_sd.get_value()),
            "configurationItemId": utils.extract_id(self.widget_ci.get_value()),
            "applicationModuleIds": "|".join([utils.extract_id(x) for x in self.widget_app_modules.get_selected_values()]),
            "description": self.entry_desc.value.strip(),
            "bonificaTeam": self.widget_bonifica_team.get_value(),     # Aggiunto
            "bonificaOffice": self.widget_bonifica_office.get_value()  # Aggiunto
        }

    def mark_form_dirty(self, e=None):
        self.form_validated = False
        self.btn_submit.disabled = True
        self.btn_submit.style = ft.ButtonStyle(bgcolor=ft.Colors.GREY_700, color=ft.Colors.WHITE, padding=20)
        self.txt_status.value = "Modifiche rilevate. Premi 'Valida associazione'."
        self.txt_status.color = ft.Colors.AMBER_300
        self.update()

    def validate_single(self, e):
        snap = self.get_single_snapshot()
        if not snap["solutionDesignId"] or not snap["configurationItemId"]:
            self.show_dialog("Validazione Fallita", "I campi 'Solution Design' e 'Configuration Item' sono obbligatori.", is_error=True)
            return
        
        self.form_validated = True
        self.btn_submit.disabled = False
        self.btn_submit.style = ft.ButtonStyle(bgcolor=ft.Colors.GREEN_700, color=ft.Colors.WHITE, padding=20)
        self.txt_status.value = "Validazione OK. Pronto per l'associazione."
        self.txt_status.color = ft.Colors.GREEN_300
        self.update()

    def submit_single(self, e):
        api_key = self.app.config.get("api_key", "")
        if not api_key:
            self.show_dialog("Errore", "API Key mancante nelle impostazioni.", is_error=True)
            return

        self.btn_submit.disabled = True
        self.btn_submit.text = "Elaborazione..."
        self.update()

        try:
            res = execute_association_api(api_key, self.get_single_snapshot())
            if res["success"]:
                self.show_dialog("Successo", res["message"])
                self.app.config.setdefault("stats", {})["ci_singoli_ok"] = self.app.config.get("stats", {}).get("ci_singoli_ok", 0) + 1
                
                # Nascondi di nuovo i campi bonifica in caso di successo
                self.widget_bonifica_team.visible = False
                self.widget_bonifica_office.visible = False
            else:
                # Intercettazione errore per Bonifica
                err_str = " ".join(res["errors"]).lower()
                if "team" in err_str or "uffici" in err_str or "obbligatorio" in err_str:
                    self.widget_bonifica_team.visible = True
                    self.widget_bonifica_office.visible = True
                    self.show_dialog(
                        "🛠️ Bonifica Necessaria", 
                        "Il CI selezionato è legacy e manca dei Team/Uffici obbligatori.\nSono comparsi dei nuovi campi nel form: compilali per forzare la bonifica e premi di nuovo 'Associa CI'.", 
                        is_error=True
                    )
                else:
                    self.show_dialog("Errore API", "\n".join(res["errors"]), is_error=True)
                
                self.app.config.setdefault("stats", {})["ci_ko"] = self.app.config.get("stats", {}).get("ci_ko", 0) + 1
        except Exception as ex:
            self.show_dialog("Errore Critico", str(ex), is_error=True)
        finally:
            config.save_app_config(self.app.config)
            self.btn_submit.text = "Associa CI"
            self.mark_form_dirty()

    def download_template(self, e):
        root = tk.Tk()
        root.attributes("-topmost", True)
        root.withdraw()
        path = filedialog.asksaveasfilename(title="Salva Template Excel Associazione", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], initialfile="Template_Associazione_CI.xlsx")
        root.destroy()
        
        if path:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(api_queries.EXCEL_ASSOC_HEADER)
            
            # Se la UI ha dati compilati, li pre-popoliamo nel template
            snap = self.get_single_snapshot()
            if snap["configurationItemId"] or snap["solutionDesignId"]:
                ws.append([snap["configurationItemId"], snap["solutionDesignId"], snap["applicationModuleIds"], snap["description"]])
            
            wb.save(path)
            self.show_dialog("Template Creato", f"Template Excel salvato in:\n{path}")

    # =========================================================================
    # LOGICA MASSIVA (CON PREVIEW MAPPER)
    # =========================================================================
    def pick_file_native(self, e):
        root = tk.Tk()
        root.attributes('-topmost', True)
        root.withdraw()
        file_path = filedialog.askopenfilename(title="Seleziona Template Excel", filetypes=[("Excel files", "*.xlsx")])
        root.destroy()

        if file_path:
            self.massivo_filepath = file_path
            self.txt_massivo_file.value = os.path.basename(file_path)
            self.txt_massivo_file.color = ft.Colors.GREEN_300
            self.btn_massivo_start.disabled = False
            self.btn_massivo_preview.disabled = False
            self.app.log(f"[MASSIVO-ASSOC] File caricato: {file_path}")
        self.update()

    def _map_id_to_logical(self, id_str: str, db_key: str) -> str:
        if not id_str: return ""
        db = self.app.local_db.get(db_key, {})
        ids = [i.strip() for i in str(id_str).split('|') if i.strip()]
        return " | ".join([db.get(i, f"ID:{i}") for i in ids])

    def show_preview_dialog(self, e):
        if not self.massivo_filepath: return
        try:
            wb = openpyxl.load_workbook(self.massivo_filepath, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            if "configurationItemId" not in headers:
                self.show_dialog("Errore Anteprima", "Il file Excel non contiene 'configurationItemId'.", is_error=True)
                return

            columns = [ft.DataColumn(ft.Text(h, weight=ft.FontWeight.BOLD, color=ft.Colors.BLUE_200)) for h in headers]
            rows = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if not row_dict.get("configurationItemId"): continue
                
                cells = []
                for h in headers:
                    raw_val = str(row_dict.get(h) or "").strip()
                    if h == "configurationItemId": logical_val = self._map_id_to_logical(raw_val, "configuration_items")
                    elif h == "solutionDesignId": logical_val = self._map_id_to_logical(raw_val, "solution_designs")
                    elif h == "applicationModuleIds": logical_val = self._map_id_to_logical(raw_val, "app_modules")
                    else: logical_val = raw_val
                    cells.append(ft.DataCell(ft.Text(logical_val, size=12)))
                rows.append(ft.DataRow(cells=cells))

            table = ft.DataTable(columns=columns, rows=rows, heading_row_color=ft.Colors.GREY_900, border=ft.border.all(1, ft.Colors.GREY_800))
            dlg = ft.AlertDialog(
                title=ft.Row([ft.Icon(ft.Icons.TABLE_CHART, color=ft.Colors.BLUE), ft.Text("Anteprima Dati Associazione")]),
                content=ft.Container(content=ft.Column([ft.Row([table], scroll=ft.ScrollMode.ALWAYS)], scroll=ft.ScrollMode.ALWAYS), width=1000, height=500, border_radius=10),
                actions=[ft.TextButton("Chiudi", on_click=lambda e: self.close_dialog(dlg))]
            )
            self.app.page.show_dialog(dlg)
        except Exception as ex:
            self.show_dialog("Errore Anteprima", str(ex), is_error=True)

    def start_massive(self, e):
        if not self.massivo_filepath: return
        api_key = self.app.config.get("api_key", "")
        if not api_key:
            self.app.log("ERRORE: API Key mancante.", level="ERROR")
            return

        self.btn_massivo_start.disabled = True
        self.btn_massivo_preview.disabled = True
        self.btn_massivo_pick.disabled = True
        self.massivo_progress.value = 0
        self.massivo_progress.visible = True
        self.txt_massivo_status.value = "Elaborazione associazioni in corso..."
        self.update()

        threading.Thread(target=self._process_massivo_thread, args=(api_key,), daemon=True).start()

    def _process_massivo_thread(self, api_key):
        try:
            self.massivo_failed_rows = [] # Resetta gli errori a ogni nuova esecuzione
            
            wb = openpyxl.load_workbook(self.massivo_filepath, data_only=True)
            sheet = wb.active
            headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
            
            valid_rows = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                if row_dict.get("configurationItemId"):
                    valid_rows.append({k: str(v).strip() if v is not None else "" for k, v in row_dict.items() if k})

            total = len(valid_rows)
            success_count = 0
            error_count = 0

            for i, snapshot in enumerate(valid_rows):
                ci_id = snapshot.get("configurationItemId", "?")
                self.app.log(f"[MASSIVO-ASSOC] Invio CI ID {ci_id}...")
                
                res = execute_association_api(api_key, snapshot)
                
                if res.get("success"):
                    self.app.log(f" ✅ OK: CI ID {ci_id} associato.")
                    success_count += 1
                else:
                    err_msg = " | ".join(res.get("errors", []))
                    self.app.log(f" ❌ KO: CI ID {ci_id} - {res.get('message')} | {err_msg}", level="ERROR")
                    error_count += 1
                    
                    # Salviamo lo snapshot originale fallito con il motivo dell'errore
                    snapshot["MOTIVO_ERRORE"] = err_msg
                    self.massivo_failed_rows.append(snapshot)

                self.massivo_progress.value = (i + 1) / total
                self.app.page.update()

            self.app.config.setdefault("stats", {})["ci_massivi_ok"] = self.app.config.get("stats", {}).get("ci_massivi_ok", 0) + success_count
            self.app.config.setdefault("stats", {})["ci_ko"] = self.app.config.get("stats", {}).get("ci_ko", 0) + error_count
            config.save_app_config(self.app.config)

            msg = f"Completato. Successi: {success_count}, Errori: {error_count}"
            self.app.log(f"[MASSIVO-ASSOC] {msg}")
            
            # Mostra il bottone del report se ci sono stati errori
            if error_count > 0:
                self.btn_massivo_report.visible = True
                
            self._unlock_massivo_ui(msg)

        except Exception as e:
            self.app.log(f"ERRORE THREAD MASSIVO: {str(e)}", level="ERROR")
            self._unlock_massivo_ui("Errore di sistema.")

    def _unlock_massivo_ui(self, msg):
        self.btn_massivo_start.disabled = False
        self.btn_massivo_preview.disabled = False
        self.btn_massivo_pick.disabled = False
        self.massivo_progress.visible = False
        self.txt_massivo_status.value = msg
        self.app.page.update()

    # =========================================================================
    # DIALOG GLOBALE
    # =========================================================================
    # ... fine di _unlock_massivo_ui ...
    def _unlock_massivo_ui(self, msg):
        self.btn_massivo_start.disabled = False
        self.btn_massivo_preview.disabled = False
        self.btn_massivo_pick.disabled = False
        self.massivo_progress.visible = False
        self.txt_massivo_status.value = msg
        self.app.page.update()

    # =========================================================================
    # NUOVO METODO: EXPORT REPORT ERRORI (INCOLLA QUESTO)
    # =========================================================================
    def download_error_report(self, e):
        """Genera un file Excel contenente solo i record andati in KO con il dettaglio dell'errore."""
        if not self.massivo_failed_rows:
            self.show_dialog("Info", "Non ci sono errori da esportare.")
            return

        root = tk.Tk()
        root.attributes("-topmost", True)
        root.withdraw()
        path = filedialog.asksaveasfilename(
            title="Salva Report Errori Massivo", 
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx")], 
            initialfile="Report_KO_Associazioni.xlsx"
        )
        root.destroy()
        
        if path:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                
                # Ricaviamo gli header dinamicamente dalle chiavi del primo dizionario scartato
                headers = list(self.massivo_failed_rows[0].keys())
                ws.append(headers)
                
                # Inseriamo i dati
                for row_dict in self.massivo_failed_rows:
                    row_data = [row_dict.get(h, "") for h in headers]
                    ws.append(row_data)
                
                wb.save(path)
                self.show_dialog("Report Generato", f"Il file con gli scarti è stato salvato in:\n{path}\n\nPuoi correggere i dati in questo file e ricaricarlo direttamente nel massivo.")
                self.app.log(f"[MASSIVO-ASSOC] Report errori esportato: {path}")
            except Exception as ex:
                self.show_dialog("Errore Export", str(ex), is_error=True)
                self.app.log(f"[MASSIVO-ASSOC] Errore export report: {str(ex)}", level="ERROR")

    #
    def show_dialog(self, title, content, is_error=False):
        dlg = ft.AlertDialog(
            title=ft.Row([ft.Icon(ft.Icons.INFO_OUTLINE, color=ft.Colors.RED_ACCENT if is_error else ft.Colors.BLUE_400), ft.Text(title, weight=ft.FontWeight.BOLD)]),
            content=ft.Column([ft.Text(content, size=14)], scroll=ft.ScrollMode.AUTO, height=200 if len(content) > 150 else None, tight=True),
            actions=[ft.TextButton("Chiudi", on_click=lambda e: self.close_dialog(dlg))],
            shape=ft.RoundedRectangleBorder(radius=10), bgcolor=ft.Colors.GREY_900
        )
        self.app.page.show_dialog(dlg)

    def close_dialog(self, dlg):
        if hasattr(self.app.page, "close_dialog"): self.app.page.close_dialog()
        elif hasattr(self.app.page, "pop_dialog"): self.app.page.pop_dialog()
        else: dlg.open = False; self.app.page.update()